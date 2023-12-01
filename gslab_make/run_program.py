# -*- coding: utf-8 -*-
from builtins import (bytes, str, open, super, range,
                      zip, round, input, int, pow, object)

import os
import re
import sys
import shutil
import traceback
import fileinput
import subprocess
import platform
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

if platform.system() == 'Windows':
    import win32com.client

import nbformat
from nbconvert.preprocessors import ExecutePreprocessor

from termcolor import colored
import colorama
colorama.init()

import gslab_make.private.messages as messages
import gslab_make.private.metadata as metadata
from gslab_make.private.exceptionclasses import CritError, ColoredError, ProgramError
from gslab_make.private.programdirective import Directive, ProgramDirective, SASDirective, LyXDirective
from gslab_make.private.utility import get_path, format_message, norm_path
from gslab_make.write_logs import write_to_makelog


def run_jupyter(paths, program, timeout = None, kernel_name = ''):
    """.. Run Jupyter notebook using system command.

    Runs notebook ``program`` using Python API, with notebook specified 
    in the form of ``notebook.ipynb``. 
    Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    timeout : int, optional
        Time to wait (in seconds) to finish executing a cell before raising exception. 
        Defaults to no timeout.
    kernel_name : str, optional
        Name of kernel to use for execution 
        (e.g., ``python2`` for standard Python 2 kernel, ``python3`` for standard Python 3 kernel). 
        Defaults to ``''`` (i.e., kernel specified in notebook).

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_jupyter(paths, program = 'notebook.ipynb')
    """

    try:
        program = norm_path(program)

        with open(program) as f:
            message = 'Processing notebook: `%s`' % program
            write_to_makelog(paths, message)    
            print(colored(message, 'cyan'))
            
            if not kernel_name:
                kernel_name = 'python%s' % sys.version_info[0]
            ep = ExecutePreprocessor(timeout = timeout, kernel_name = kernel_name)
            nb = nbformat.read(f, as_version = 4)       
            ep.preprocess(nb, {'metadata': {'path': '.'}})
            
        with open(program, 'wt') as f:
            nbformat.write(nb, f)
    except:
        error_message = 'Error with `run_jupyter`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_lyx(paths, program, doctype = '', **kwargs): 
    """.. Run LyX script using system command.

    Compiles document ``program`` using system command, with document specified 
    in the form of ``script.lyx``. Status messages are appended to file ``makelog``. 
    PDF outputs are written in directory ``output_dir``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.
    doctype : str, optional
        Type of LyX document. Takes either ``'handout'`` and ``'comments'``. 
        All other strings will default to standard document type. 
        Defaults to ``''`` (i.e., standard document type).

    Path Keys
    ---------
    makelog : str
        Path of makelog.
    output_dir : str
        Directory to write PDFs.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_lyx(paths, program = 'script.lyx')
    """

    try:
        makelog = get_path(paths, 'makelog')
        output_dir = get_path(paths, 'output_dir')
        direct = LyXDirective(output_dir = output_dir, 
                              doctype = doctype,
                              application = 'lyx', 
                              program = program, 
                              makelog = makelog, 
                              **kwargs)
            
        # Make handout/comments LyX file        
        if direct.doctype:
            temp_name = os.path.join(direct.program_name + '_' + direct.doctype)
            temp_program = os.path.join(direct.program_dir, temp_name + '.lyx') 
            
            beamer = False
            shutil.copy2(direct.program, temp_program) 

            for line in fileinput.input(temp_program, inplace = True, backup = '.bak'):
                if r'\textclass beamer' in line:
                    beamer = True          
                if direct.doctype == 'handout' and beamer and (r'\options' in line):
                    line = line.rstrip('\n') + ', handout\n'
                elif direct.doctype == 'comments' and (r'\begin_inset Note Note' in line):
                    line = line.replace('Note Note', 'Note Greyedout')
                
                print(line)
        else:
            temp_name = direct.program_name
            temp_program = direct.program

        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.option, temp_program)
        exit_code, stderr = direct.execute_command(command)
        direct.write_log()
        if exit_code != 0:
            error_message = 'LyX program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)

        # Move PDF output
        temp_pdf = os.path.join(direct.program_dir, temp_name + '.pdf')
        output_pdf = os.path.join(direct.output_dir, direct.program_name + '.pdf')

        if temp_pdf != output_pdf:
            shutil.copy2(temp_pdf, output_pdf)
            os.remove(temp_pdf)
            
        # Remove handout/comments LyX file
        if direct.doctype:
            os.remove(temp_program)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_lyx`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_latex(paths, program, **kwargs): 
    """.. Run LaTeX script using system command.

    Compiles document ``program`` using system command, with document specified 
    in the form of ``script.tex``. Status messages are appended to file ``makelog``. 
    PDF outputs are written in directory ``output_dir``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.
    output_dir : str
        Directory to write PDFs.

    Note
    ----
    We recommend leaving all other parameters to their defaults.
    
    Note
    ----
    This function creates and removes a directory named ``latex_auxiliary_dir``.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_latex(paths, program = 'script.tex')
    """

    try:
        makelog = get_path(paths, 'makelog')
        output_dir = get_path(paths, 'output_dir')
        direct = LyXDirective(output_dir = output_dir, 
                              application = 'latex', 
                              program = program, 
                              makelog = makelog, 
                              **kwargs)
            
        temp_name = direct.program_name
        temp_program = direct.program

        # Generate folder for auxiliary files
        if os.path.exists('latex_auxiliary_dir'):
            # If it exists, remove the directory and its contents
            shutil.rmtree('latex_auxiliary_dir')
        os.mkdir('latex_auxiliary_dir')
        
        # Shift path if necessary
        original_dir = os.getcwd()
        program_dir = os.path.dirname(program)
        if program_dir:
            os.chdir(program_dir)
        temp_depth = len([p for p in program_dir.split(os.path.sep) if p])
        back_original_dir = os.path.sep.join(['..'] * temp_depth) + os.path.sep if temp_depth else ''

        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, back_original_dir, direct.option, temp_program)
        exit_code, stderr = direct.execute_command(command)
        os.chdir(original_dir)
        direct.write_log()
        if exit_code != 0:
            error_message = 'LaTeX program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)

        # Move PDF output
        temp_pdf = os.path.join('latex_auxiliary_dir', temp_name + '.pdf')
        output_pdf = os.path.join(direct.output_dir, direct.program_name + '.pdf')

        if temp_pdf != output_pdf:
            shutil.copy2(temp_pdf, output_pdf)
            shutil.rmtree('latex_auxiliary_dir')
        
        # Remove auxiliary files
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_latex`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_mathematica(paths, program, **kwargs):
    """.. Run Mathematica script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.m``. Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_mathematica(paths, program = 'script.m')
    """
    
    try:
        makelog = get_path(paths, 'makelog')
        direct = ProgramDirective(application = 'math', 
                                  program = program, 
                                  makelog = makelog, 
                                  **kwargs)

        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.program, direct.option)
        exit_code, stderr = direct.execute_command(command)
        direct.write_log()
        if exit_code != 0:
            error_message = 'Mathematica program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_mathematica`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_matlab(paths, program, **kwargs):
    """.. Run Matlab script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.m``. Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_matlab(paths, program = 'script.m')
    """

    try:
        makelog = get_path(paths, 'makelog')
        direct = ProgramDirective(application = 'matlab', 
                                  program = program, 
                                  makelog = makelog, 
                                  **kwargs)
        
        # Get program output
        program_log = os.path.join(os.getcwd(), direct.program_name + '.log')

        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.option, direct.program, direct.program_name + '.log')
        exit_code, stderr = direct.execute_command(command)   
        if exit_code != 0:
            error_message = 'Matlab program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
        direct.move_program_output(program_log, direct.log)   
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_matlab`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_perl(paths, program, **kwargs):
    """.. Run Perl script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.pl``. Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Arguments for system command. Defaults to no arguments.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_perl(paths, program = 'script.pl')
    """

    try:
        makelog = get_path(paths, 'makelog')
        direct = ProgramDirective(application = 'perl', 
                                  program = program, 
                                  makelog = makelog, 
                                  **kwargs)
        
        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.option, direct.program, direct.args)
        exit_code, stderr = direct.execute_command(command)
        direct.write_log()
        if exit_code != 0:
            error_message = 'Perl program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_perl`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_python(paths, program, **kwargs):
    """.. Run Python script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.py``. Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Arguments for system command. Defaults to no arguments.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_python(paths, program = 'script.py')
    """

    try:
        makelog = get_path(paths, 'makelog')
        direct = ProgramDirective(application = 'python', 
                                  program = program, 
                                  makelog = makelog, 
                                  **kwargs)

        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.option, direct.program, direct.args)
        exit_code, stderr = direct.execute_command(command)
        direct.write_log() 
        if exit_code != 0:
            error_message = 'Python program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_python`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_r(paths, program, **kwargs):
    """.. Run R script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.R``. Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_r(paths, program = 'script.R')
    """
    
    try:
        makelog = get_path(paths, 'makelog')
        direct = ProgramDirective(application = 'r', 
                                  program = program, 
                                  makelog = makelog, 
                                  **kwargs)

        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.option, direct.program)
        exit_code, stderr = direct.execute_command(command)
        direct.write_log()      
        if exit_code != 0:
            error_message = 'R program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_r`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())

def run_julia(paths, program, **kwargs):
    """.. Run Julia script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.jl``. Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_julia(paths, program = 'script.jl')
    """
    
    try:
        makelog = get_path(paths, 'makelog')
        direct = ProgramDirective(application = 'julia', 
                                  program = program, 
                                  makelog = makelog, 
                                  **kwargs)

        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.option, direct.program)
        exit_code, stderr = direct.execute_command(command)
        direct.write_log()      
        if exit_code != 0:
            error_message = 'Julia program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_julia`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_sas(paths, program, lst = '', **kwargs):
    """.. Run SAS script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.sas``. Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.
    lst : str, optional
        Path of program lst. Program lst is only written if specified. 
        Defaults to ``''`` (i.e., not written).

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_sas(paths, program = 'script.sas')
    """

    try:
        makelog = get_path(paths, 'makelog')
        direct = SASDirective(application = 'sas', 
                              program = program, 
                              makelog = makelog, 
                              **kwargs)

        # Get program outputs
        program_log = os.path.join(os.getcwd(), direct.program_name + '.log')
        program_lst = os.path.join(os.getcwd(), direct.program_name + '.lst')
        
        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.option, direct.program)       
        exit_code, stderr = direct.execute_command(command)
        if exit_code != 0:
            error_message = 'SAS program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
        direct.move_program_output(program_log)
        direct.move_program_output(program_lst)        
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_sas`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_stat_transfer(paths, program, **kwargs):
    """.. Run StatTransfer script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.stc`` or ``script.stcmd``. 
    Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_stat_transfer(paths, program = 'script.stc')
    """

    try:
        makelog = get_path(paths, 'makelog')
        direct = ProgramDirective(application = 'st', 
                                  program = program, 
                                  makelog = makelog, 
                                  **kwargs)

        # Execute
        command = metadata.commands[direct.osname][direct.application] % (direct.executable, direct.program)
        exit_code, stderr = direct.execute_command(command)
        direct.write_log()
        if exit_code != 0:
            error_message = 'StatTransfer program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_stat_transfer`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_stata(paths, program, **kwargs):
    """.. Run Stata script using system command.

    Runs script ``program`` using system command, with script specified 
    in the form of ``script.do``. Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    program : str
        Path of script to run.

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Note
    ----
    When a do-file contains a space in its name, different version of Stata save the
    corresponding log file with different names. Some versions of Stata truncate the 
    name to everything before the first space of the do-file name.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of program log. Program log is only written if specified. 
        Defaults to ``''`` (i.e., not written). 
    executable : str, optional
        Executable to use for system command. 
        Defaults to executable specified in :ref:`default settings<default settings>`.
    option : str, optional
        Options for system command. Defaults to options specified in :ref:`default settings<default settings>`.
    args : str, optional
        Not applicable.

    Returns
    -------
    None

    Example
    -------
    .. code-block:: python

        run_stata(paths, program = 'script.do')
    """

    try:
        makelog = get_path(paths, 'makelog')
        direct = ProgramDirective(application = 'stata', 
                                  program = program, 
                                  makelog = makelog, 
                                  **kwargs)

        # Get program output (partial)
        program_name = direct.program.split(" ")[0]
        program_name = os.path.split(program_name)[-1]
        program_name = os.path.splitext(program_name)[0]
        program_log_partial = os.path.join(os.getcwd(), program_name + '.log')
        
        # Get program output (full)
        program_log_full = os.path.join(os.getcwd(), direct.program_name + '.log')

        # Sanitize program 
        if direct.osname == "posix":
            direct.program = re.escape(direct.program)

        # Execute
        command = metadata.commands[direct.osname]['stata'] % (direct.executable, direct.option, direct.program)
        exit_code, stderr = direct.execute_command(command)
        if exit_code != 0:
            error_message = 'Stata program executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
        try:
            output = direct.move_program_output(program_log_partial, direct.log)
        except:
            output = direct.move_program_output(program_log_full, direct.log)
        _check_stata_output(output)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `run_stata`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())
    
    
def _check_stata_output(output):
    """.. Check Stata output"""
    
    regex = "end of do-file[\s]*r\([0-9]*\);"
    if re.search(regex, output):
        error_message = 'Stata program executed with errors.'
        error_message = format_message(error_message)
        raise ProgramError(error_message, 'See makelog for more detail.')
    
def install_pdf_crop_margins():
    """.. Check if pdf-crop-margins is installed and install it if it's not."""
    try:
        subprocess.run(["pdf-crop-margins", "--version"], stdout = subprocess.DEVNULL, stderr = subprocess.DEVNULL)
    except subprocess.CalledProcessError as e:
        print("pdf-crop-margins is not installed. Please install it.")
    except FileNotFoundError:
        print("pdf-crop-margins not found. Installing it now...")
        try:
            subprocess.run(["pip", "install", "pdfCropMargins"], stdout = subprocess.DEVNULL, stderr = subprocess.DEVNULL)
            print("pdf-crop-margins has been successfully installed.")
        except subprocess.CalledProcessError as e:
            print("An error occurred while trying to install pdf-crop-margins. Please install it manually.")

def write_excel_scalars(template, scalar):

    if not scalar:
        return

    if not os.path.exists(template):
        raise FileNotFoundError(f"Template file '{template}' not found.")

    if not os.path.exists(scalar):
        raise FileNotFoundError(f"Scalar file '{scalar}' not found.")

    try:
        # Load the scalar data.
        scalar_df = pd.read_excel(scalar, sheet_name = 0, header = None)

        # Load the template workbook.
        book = openpyxl.load_workbook(template)

        # Remove the first sheet if it exists.
        first_sheet_name = book.sheetnames[0]
        if first_sheet_name in book.sheetnames:
            del book[first_sheet_name]

        # Create a new sheet at the first position.
        sheet = book.create_sheet(first_sheet_name, 0)

        # Write DataFrame to the new sheet, starting from the first row.
        for r_idx, row in enumerate(scalar_df.values, start = 1):
            for c_idx, value in enumerate(row, start = 1):
                sheet.cell(row = r_idx, column = c_idx, value = value)

        # Save and close the workbook.
        book.save(template)
        book.close()

    except Exception as e:
        raise RuntimeError(f"Error in `write_excel_scalars`: {e}")
    
def export_excel_tables(paths, template, scalar, **kwargs):
    """.. Convert Excel template file to PDF using Microsoft Excel's native functionality.
    
    Converts Excel document specified by `template` to a PDF file. The resulting PDF
    will be saved in the `output_dir` specified within the `paths` dictionary.
    
    ------------------------------
    Instructions for first-time usage on macOS:
    ------------------------------

    For Mac Users: When first running this script, you must grant Microsoft Excel access to the folder
    where the Excel files are located and where the PDFs will be saved (i.e. your local
    repository). This is a one-time setup to allow the script to run without interruptions. You will
    be notified by a pop-up window whether to grant these permissions. If you select "Yes", you can
    check the status of permissions with the following steps:
    
    1. Open 'System Settings' from the Apple menu or search for it using Spotlight.
    
    2. Go to the 'Privacy & Security' tab.
    
    3. From the list on the right, choose 'Files and Folders', then navigate to the Microsoft Excel icon.
    
    4. In the associated dropdown, you should see the folders that you have granted Microsoft Excel access to.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Should contain 'makelog' and 'output_dir' keys with
        corresponding values.
    template : str
        Name of the Excel file to convert, expected to be in the `output_dir`.
    scalar : str
        Name of the scalar sheet to fill with the populated values from ~/analysis/output.
    
    Other Parameters
    ----------------
    osname : str, optional
        Name of the operating system. Used to determine the method of conversion.
        Defaults to the system's actual OS name.
    shell : bool, optional
        If using subprocess, determines whether to use the shell.
        Defaults to False.
    log : str, optional
        Path to a log file where the function should append status messages.
    
    Returns
    -------
    None
    
    Example
    -------
    .. code-block:: python
    
        paths = {
            'makelog': 'path/to/makelog.txt',
            'output_dir': 'path/to/output'
        }
        
        export_excel_tables(PATHS, template = 'tables/skeletons/example_table.xlsx', scalar = 'tables/scalars/example_scalar.xlsx')
    
    """
    
    try:

       # Install PDF crop margins if not available.
        install_pdf_crop_margins()

        # Extract the relevant paths
        makelog = paths.get('makelog', '')

        # Get the directory of the calling script
        script_caller_dir = os.getcwd()  # This gets the current working directory

        # Calculate the paths relative to the script location
        output_dir = os.path.join(script_caller_dir, 'output')

        # Construct the full paths to the files
        skeleton_file_path = os.path.join(script_caller_dir, 'tables/skeletons', template + '.xlsx')
        skeleton_file_path_modified = os.path.join(script_caller_dir, 'tables/skeletons', template + '_modified.xlsx')
        pdf_file_name = os.path.splitext(template)[0] + '.pdf'
        pdf_output_path = os.path.join(output_dir, pdf_file_name)

        # Populates template skeleton with scalar values.
        if scalar != False:
            scalar_file_path = os.path.join(script_caller_dir, 'input/tables', scalar)
            write_to_makelog(paths, scalar_file_path)
            write_excel_scalars(skeleton_file_path, scalar_file_path)

        # Determine the operating system
        osname = platform.system()
        shell = kwargs.get('shell', False)

        # Excel to PDF conversion
        if osname == 'Darwin':  # macOS

            # Convert to POSIX path format
            posix_skeleton_file_path = skeleton_file_path.replace(os.sep, '/')
            posix_pdf_output_path = pdf_output_path.replace(os.sep, '/')
            posix_skeleton_file_path_modified = skeleton_file_path_modified.replace(os.sep, '/')

            write_to_makelog(paths, posix_pdf_output_path)
            write_to_makelog(paths, posix_skeleton_file_path)
            write_to_makelog(paths, posix_skeleton_file_path_modified)

            # Prepare the AppleScript command
            applescript_command = f'''

            set skeleton_file_path to "{posix_skeleton_file_path}"
            set pdf_output_path to "{posix_pdf_output_path}"
            set skeleton_file_path_modified to "{posix_skeleton_file_path_modified}"

            tell application "Microsoft Excel"
                
                activate

                -- Suppress alerts to avoid confirmation dialogs
                set display alerts to false
                
                -- Open the initial workbook
                open skeleton_file_path

                copy worksheet worksheet "Table"
                activate object
                set theSheet to sheet 1 of active workbook
                
                -- Set margins.
                tell page setup object of theSheet
                    set page orientation to landscape
                    set zoom to false
                    set fit to pages wide to 1
                    set fit to pages tall to 9999
                end tell
                
                -- Export as a PDF.
                save as theSheet filename pdf_output_path file format PDF file format

                -- Close the workbooks without saving.
                close active workbook saving no

            end tell

            '''

            # Execute the AppleScript command
            process = subprocess.run(["osascript", "-e", applescript_command], capture_output = True, text = True, shell = shell)

            if process.returncode != 0:
                print(f"AppleScript Error: {process.stderr}")
                raise Exception(f"AppleScript Error: {process.stderr}")
            
            temp_files_path = os.path.join(script_caller_dir, 'tables/skeletons')
            for filename in os.listdir(temp_files_path):
                if filename.startswith('~$') and filename.endswith('.xlsx'):
                    os.remove(os.path.join(temp_files_path, filename))

        elif osname == 'Windows':
            # Start an instance of Excel
            excel_app = win32com.client.DispatchEx("Excel.Application")
            # Open the Excel file
            workbook = excel_app.Workbooks.Open(skeleton_file_path)
            # Select the second sheet
            worksheet = workbook.Worksheets[2]
            # Save the active sheet to a PDF
            worksheet.ExportAsFixedFormat(0, pdf_output_path)
            # Close the workbook without saving changes
            workbook.Close(SaveChanges = False)
            # Quit Excel
            excel_app.Quit()

        else:
            raise ValueError("Unsupported OS type. Function supports 'Darwin' (macOS) and 'Windows' OS names.")

        # Optionally write to a log file
        if 'log' in kwargs:
            with open(kwargs['log'], 'a') as log_file:
                log_file.write(f"Successfully converted {template} to PDF.\n")

        temp_pdf_output_path = pdf_output_path.replace('.pdf', '_temp.pdf')
        subprocess.run(["pdf-crop-margins", "-p", "0", "-a", "-6", pdf_output_path, "-o", temp_pdf_output_path])
        os.replace(temp_pdf_output_path, pdf_output_path)

    except Exception as e:
        error_message = f"Error in `export_excel_tables` for {template}: {e}\n"
        if makelog:
            with open(makelog, 'a') as makelog_file:
                makelog_file.write(error_message)
        print(error_message)
        raise RuntimeError(error_message)
    
def quit_excel(paths, **kwargs):
    """.. Quits Excel using native application following PDF exports.
    
    Parameters
    ----------
    paths : dict
        Dictionary of paths. Should contain 'makelog' and 'output_dir' keys with
        corresponding values.
    
    Other Parameters
    ----------------
    osname : str, optional
        Name of the operating system. Used to determine the method of conversion.
        Defaults to the system's actual OS name.
    shell : bool, optional
        If using subprocess, determines whether to use the shell.
        Defaults to False.
    log : str, optional
        Path to a log file where the function should append status messages.
    
    Returns
    -------
    None
    
    Example
    -------
    .. code-block:: python
    
        paths = {
            'makelog': 'path/to/makelog.txt',
            'output_dir': 'path/to/output'
        }
        
        quit_excel(PATHS)
    
    """

    try:

        # Extract the relevant paths
        makelog = paths.get('makelog', '')

        # Determine the operating system
        osname = platform.system()
        shell = kwargs.get('shell', False)

        if osname == 'Darwin':  # macOS

            # Prepare the AppleScript command
            applescript_command = f'''
            tell application "Microsoft Excel"
                quit
            end tell
            '''

            # Execute the AppleScript command
            process = subprocess.run(["osascript", "-e", applescript_command], capture_output = True, text = True, shell = shell)

            if process.returncode != 0:
                print(f"AppleScript Error: {process.stderr}")
                raise Exception(f"AppleScript Error: {process.stderr}")
            
    except Exception as e:
        error_message = f"Error in `export_excel_tables` for {template}: {e}\n"
        if makelog:
            with open(makelog, 'a') as makelog_file:
                makelog_file.write(error_message)
        print(error_message)
        raise RuntimeError(error_message)

def execute_command(paths, command, **kwargs):
    """.. Run system command.

    Runs system command `command` with shell execution boolean ``shell``. 
    Outputs are appended to file ``makelog`` and written to system command log file ``log``. 
    Status messages are appended to file ``makelog``.

    Parameters
    ----------
    paths : dict
        Dictionary of paths. Dictionary should contain values for all keys listed below.
    command : str
        System command to run.
    shell : `bool`, optional
        See `here <https://docs.python.org/3/library/subprocess.html#frequently-used-arguments>`_. 
        Defaults to ``True``.
    log : str, optional
        Path of system command log. System command log is only written if specified. 
        Defaults to ``''`` (i.e., not written).

    Path Keys
    ---------
    makelog : str
        Path of makelog.

    Note
    ----
    We recommend leaving all other parameters to their defaults.

    Other Parameters
    ----------------
    osname : str, optional
        Name of OS. Used to check if OS is supported. Defaults to ``os.name``.


    Returns
    -------
    None

    Example
    -------
    The following code executes the ``ls`` command, 
    writes outputs to system command log file ``'file'``, 
    and appends outputs and/or status messages to ``paths['makelog']``.

    .. code-block:: python

        execute_command(paths, 'ls', log = 'file')
    """
    
    try:
        makelog = get_path(paths, 'makelog')
        direct = Directive(makelog = makelog, **kwargs)

        # Execute
        exit_code, stderr = direct.execute_command(command)
        direct.write_log()   
        if exit_code != 0:
            error_message = 'Command executed with errors. Traceback can be found below.'
            error_message = format_message(error_message)
            raise ProgramError(error_message, stderr)
    except ProgramError:
        raise
    except:
        error_message = 'Error with `execute_command`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        write_to_makelog(paths, error_message + '\n\n' + traceback.format_exc())
        raise ColoredError(error_message, traceback.format_exc())


def run_module(root, module, build_script = 'make.py', osname = None, run_all=True):
    """.. Run module. 
    
    Runs script `build_script` in module directory `module` relative to root of repository `root`.

    Parameters
    ----------
    root : str 
        Directory of root.
    module: str
        Name of module.
    build_script : str
        Name of build script. Defaults to ``make.py``.
    osname : str, optional
        Name of OS. Used to determine syntax of system command. Defaults to ``os.name``.
    run_all : bool
        If being run from the root. Will make it so doesn't recheck conda status.

    Returns
    -------
    None
    
    Example
    -------
    The following code runs the script ``root/module/make.py``.

    .. code-block:: python

        run_module(root = 'root', module = 'module')
    """

    osname = osname if osname else os.name # https://github.com/sphinx-doc/sphinx/issues/759

    try:
        module_dir = os.path.join(root, module)
        os.chdir(module_dir)

        build_script = norm_path(build_script)
        if not os.path.isfile(build_script):
            raise CritError(messages.crit_error_no_file % build_script)  

        message = 'Running module `%s`' % module
        message = format_message(message)
        message = colored(message, attrs = ['bold'])
        print('\n' + message)  

        if run_all:
            status = os.system('%s %s run_all' % (metadata.default_executables[osname]['python'], build_script))
            os.chdir(os.path.normpath(os.getcwd() + os.sep + os.pardir))

        else: 
            status = os.system('%s %s' % (metadata.default_executables[osname]['python'], build_script))

        if status != 0:
            raise ProgramError()
    except ProgramError:
        sys.exit()
    except:
        error_message = 'Error with `run_module`. Traceback can be found below.' 
        error_message = format_message(error_message) 
        raise ColoredError(error_message, traceback.format_exc())


__all__ = ['run_stata', 'run_matlab', 'run_perl', 'run_python', 
           'run_jupyter', 'run_mathematica', 'run_stat_transfer', 
           'run_lyx', 'run_latex', 'run_r', 'run_sas', 'export_excel_tables',
           'quit_excel', 'execute_command', 'run_module']
